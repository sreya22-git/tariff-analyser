from __future__ import annotations
import csv
import io
import json
import re
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook


def build_sheets_payload(raw: bytes, filename: str) -> List[Dict[str, Any]]:
    """Parse raw file bytes (XLSX or CSV) into headers + a TSV sample per sheet, for LLM input."""
    ext = (filename.split(".")[-1] if "." in filename else "").lower()
    sheets: List[Dict[str, Any]] = []

    if ext == "xlsx":
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            headers = [str(h) if h is not None else "" for h in (rows[0] if rows else [])]
            sample_rows = rows[1:]
            tsv = "\n".join(["\t".join([str(v) if v is not None else "" for v in r]) for r in sample_rows])
            sheets.append({"name": ws.title, "headers": headers, "sample": tsv})
    else:
        try:
            text = raw.decode("utf-8", errors="ignore")
        except Exception:
            text = ""
        reader = list(csv.reader(io.StringIO(text)))
        headers = reader[0] if reader else []
        sample_rows = reader[1:]
        tsv = "\n".join(["\t".join(row) for row in sample_rows])
        sheets.append({"name": "Sheet1", "headers": headers, "sample": tsv})

    return sheets


def parse_llm_overview(content: str) -> Dict[str, Any]:
    try:
        return json.loads(content)
    except Exception:
        return {}


def normalize_overview(parsed: Dict[str, Any], sheets_payload: List[Dict[str, Any]]) -> Dict[str, Any]:
    result = {"sheets": []}
    for sh in parsed.get("sheets", []):
        tname = sh.get("name")
        cols_out = []
        payload = next((sp for sp in sheets_payload if sp.get("name") == tname), None)
        samples: List[List[str]] = []
        if payload and payload.get("sample"):
            samples = [row.split("\t") for row in payload["sample"].split("\n") if row]
        headers = payload.get("headers", []) if payload else []

        for c in sh.get("columns", []):
            cname = c.get("name")
            ctype = c.get("type") or c.get("dtype") or "STRING"
            desc = c.get("desc") or c.get("description") or ""
            nullable = c.get("nullable")
            if nullable is None:
                try:
                    idx = headers.index(cname) if cname in headers else None
                    if idx is not None:
                        nullable = any((r[idx].strip() == "" for r in samples if idx < len(r)))
                except Exception:
                    nullable = False
            cols_out.append({"name": cname, "type": ctype, "nullable": bool(nullable or False), "desc": desc})

        if tname:
            result["sheets"].append({"name": tname, "columns": cols_out})

    return result


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _is_fk_to(table_name: str, col_name: str) -> bool:
    t = _norm(table_name)
    t_sing = t[:-1] if t.endswith("s") else t
    c = _norm(col_name)
    return c == f"{t}id" or c == f"{t_sing}id"


_KEY_SUFFIXES = ("id", "number", "code")


def _key_prefix(col_name: str) -> Optional[str]:
    """Strip a trailing ID/Number/Code suffix to get the entity prefix a key column names,
    e.g. 'VendorID' -> 'vendor', 'PONumber' -> 'po'. Returns None if not key-shaped."""
    n = _norm(col_name)
    for suf in _KEY_SUFFIXES:
        if n.endswith(suf) and len(n) > len(suf):
            return n[: -len(suf)]
    return None


def heuristic_relationships(schema: Dict[str, Any]) -> List[Dict[str, str]]:
    """Infer relationships locally from shared column names. A shared column is treated
    as a key (ID/Number/Code suffix); the sheet whose name contains the key's entity
    prefix (e.g. 'VendorID' -> 'Vendor Master') is the parent/PK side, oriented One -> Many.
    Falls back to the single-word 'table name + id' convention (_is_fk_to) when no prefix match is found."""
    colmaps: Dict[str, Dict[str, str]] = {}
    for sh in schema.get("sheets", []):
        tname = sh.get("name", "")
        cmap = {}
        for c in sh.get("columns", []):
            if isinstance(c, dict) and c.get("name"):
                cmap[c["name"].lower()] = c["name"]
        if tname:
            colmaps[tname] = cmap

    names = list(colmaps.keys())
    rels: List[Dict[str, str]] = []
    for i, a in enumerate(names):
        for j, b in enumerate(names):
            if i == j:
                continue
            ca = colmaps.get(a, {})
            cb = colmaps.get(b, {})
            common = set(ca.keys()) & set(cb.keys())
            for col_l in common:
                col_a = ca[col_l]
                col_b = cb[col_l]
                prefix = _key_prefix(col_l)
                a_norm, b_norm = _norm(a), _norm(b)
                a_is_parent = bool(prefix) and prefix in a_norm
                b_is_parent = bool(prefix) and prefix in b_norm
                if a_is_parent and not b_is_parent:
                    rels.append({"fromTable": a, "fromColumn": col_a, "toTable": b, "toColumn": col_b, "card": "One -> Many"})
                elif b_is_parent and not a_is_parent:
                    rels.append({"fromTable": b, "fromColumn": col_b, "toTable": a, "toColumn": col_a, "card": "One -> Many"})
                elif _is_fk_to(b, col_a):
                    rels.append({"fromTable": a, "fromColumn": col_a, "toTable": b, "toColumn": col_b, "card": "Many -> One"})
                elif _is_fk_to(a, col_b):
                    rels.append({"fromTable": b, "fromColumn": col_b, "toTable": a, "toColumn": col_a, "card": "Many -> One"})

    uniq, seen = [], set()
    for r in rels:
        key = (r["fromTable"], r["fromColumn"], r["toTable"], r["toColumn"])
        if key not in seen:
            seen.add(key)
            uniq.append(r)
    return uniq


async def enrich_relationships_llm_first(
    schema: Dict[str, Any], base_url: str, api_key: str, model: str
) -> Dict[str, Any]:
    """Try LLM-inferred relationships first, fall back to the local heuristic."""
    from services.llm_service import infer_relationships

    result = dict(schema)
    rels: List[Dict[str, str]] = []
    try:
        rels = await infer_relationships(schema.get("sheets", []), base_url, api_key, model)
    except Exception:
        rels = []
    if not rels:
        rels = heuristic_relationships(schema)
    result["relationships"] = normalize_relationship_direction(schema, rels)
    return result


def normalize_relationship_direction(schema: Dict[str, Any], rels: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """Orient relationships PK -> FK with card 'One -> Many' where a PK can be identified."""
    def _pick_pk_for_table(table_name: str, columns: List[Dict[str, Any]]) -> str:
        t = _norm(table_name)
        t_sing = t[:-1] if t.endswith("s") else t
        norms = [(_norm(c.get("name")), c.get("name")) for c in columns if isinstance(c, dict) and c.get("name")]
        for cn, orig in norms:
            if cn == f"{t}id" or cn == f"{t_sing}id":
                return orig
        for cn, orig in norms:
            if cn == "id":
                return orig
        for cn, orig in norms:
            if cn.endswith("id"):
                return orig
        return ""

    pkmap: Dict[str, str] = {}
    for sh in schema.get("sheets", []):
        tname = sh.get("name", "")
        pkmap[tname] = _pick_pk_for_table(tname, sh.get("columns", []))

    out: List[Dict[str, str]] = []
    for r in rels:
        ft, fc = r.get("fromTable", ""), r.get("fromColumn", "")
        tt, tc = r.get("toTable", ""), r.get("toColumn", "")
        pk_from = pkmap.get(ft, "")
        pk_to = pkmap.get(tt, "")
        is_fk = lambda col: _norm(col).endswith("id")
        if pk_to and (not pk_from or fc != pk_from) and tc == pk_to:
            r = {"fromTable": tt, "fromColumn": pk_to, "toTable": ft, "toColumn": fc, "card": "One -> Many"}
        elif pk_from and fc == pk_from and is_fk(tc):
            r["card"] = "One -> Many"
        out.append(r)
    return out
